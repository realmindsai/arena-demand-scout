"""Fetch and parse ABS population projection XLSX files."""

import datetime
import httpx
import openpyxl
from pathlib import Path

PROJECTIONS_BASE_URL = (
    "https://www.abs.gov.au/statistics/people/population/"
    "population-projections-australia/2022-base-2071"
)

CACHE_DIR = Path(".abs_cache")

STATE_TABLE_MAP = {
    "NSW": 1, "VIC": 2, "QLD": 3, "SA": 4,
    "WA": 5, "TAS": 6, "NT": 7, "ACT": 8,
}

SERIES_MAP = {"series_a": "A", "series_b": "B", "series_c": "C"}

MAX_PROJECTION_YEAR = 2036


def download_xlsx(url: str, dest: Path) -> Path:
    """Download XLSX file if not already cached."""
    if dest.exists():
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    resp = httpx.get(url, follow_redirects=True, timeout=60)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    return dest


def download_projection_tables(cache_dir: Path | None = None) -> dict:
    """Download all projection XLSX files, return {state: {series: path}}."""
    cache = cache_dir or CACHE_DIR
    state_files = {}
    for state, num in STATE_TABLE_MAP.items():
        state_files[state] = {}
        for series_key, series_letter in SERIES_MAP.items():
            filename = f"3222_Table_{series_letter}{num}.xlsx"
            url = f"{PROJECTIONS_BASE_URL}/{filename}"
            dest = cache / filename
            download_xlsx(url, dest)
            state_files[state][series_key] = dest
    return state_files


def parse_projection_xlsx(xlsx_path: Path) -> dict:
    """Parse an ABS projection XLSX, extracting age 0-5 data by sex.

    ABS structure:
    - Row 0: Column descriptions like "Projected persons ; Series XX(Y) ; STATE ; Male ; 0 ;"
    - Rows 1-9: Metadata
    - Rows 10+: Data (col 0 = datetime, rest = population values)
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Data1" in wb.sheetnames:
        ws = wb["Data1"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row 0 has column descriptions
    headers = rows[0]

    # Build column map: identify which columns have ages 0-5 by sex
    col_map = []  # [(col_idx, sex, age)]
    for col_idx, header in enumerate(headers):
        if col_idx == 0 or not header:
            continue
        h = str(header).strip()
        h_lower = h.lower()

        sex = None
        age = None
        if "male" in h_lower and "female" not in h_lower:
            sex = "male"
        elif "female" in h_lower:
            sex = "female"

        if sex:
            parts = h.split(";")
            for part in reversed(parts):
                part = part.strip()
                if part.isdigit():
                    age = int(part)
                    break

        if sex and age is not None and 0 <= age <= 5:
            col_map.append((col_idx, sex, age))

    # Initialize data structures
    male = {age: [] for age in range(6)}
    female = {age: [] for age in range(6)}
    years = []

    # Data starts at row 10 (after 10 metadata rows: 0=headers, 1-9=metadata)
    for row in rows[10:]:
        if not row or not row[0]:
            continue
        date_val = row[0]
        if isinstance(date_val, datetime.datetime):
            year = date_val.year
        else:
            continue

        if year > MAX_PROJECTION_YEAR:
            break

        years.append(year)

        for col_idx, sex, age in col_map:
            val = row[col_idx] if col_idx < len(row) else 0
            if val is None:
                val = 0
            target = male if sex == "male" else female
            target[age].append(int(val))

    return {"years": years, "male": male, "female": female}


def aggregate_age_0_5(male: dict, female: dict) -> list[int]:
    """Sum ages 0-5 across both sexes for each year."""
    if not male or not female:
        return []
    n_years = min(len(male.get(0, [])), len(female.get(0, [])))
    totals = []
    for i in range(n_years):
        total = 0
        for age in range(6):
            total += male.get(age, [0] * n_years)[i]
            total += female.get(age, [0] * n_years)[i]
        totals.append(total)
    return totals


def build_projections_json(state_files: dict) -> dict:
    """Build the full projections JSON from downloaded XLSX files.

    state_files: {state: {series_key: Path}}

    erp_actual is derived from the projection base year (2022) Series B data.
    """
    result = {
        "base_year": 2022,
        "projection_years": [],
        "states": {},
        "erp_actual": {},
    }

    for state, series_paths in state_files.items():
        result["states"][state] = {}
        for series_key, xlsx_path in series_paths.items():
            parsed = parse_projection_xlsx(xlsx_path)
            pop_0_5 = aggregate_age_0_5(parsed["male"], parsed["female"])
            result["states"][state][series_key] = {"population_0_5": pop_0_5}

            if not result["projection_years"] and parsed["years"]:
                result["projection_years"] = parsed["years"]

            # Use Series B base year value as the ERP actual anchor
            if series_key == "series_b" and pop_0_5:
                result["erp_actual"][state] = {
                    "population_0_5": [pop_0_5[0]],
                    "as_at": "2022-06-30",
                }

    return result
